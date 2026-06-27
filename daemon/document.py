from daemon.logger import RichLogManager
from pathlib import Path

class DocumentNormalizer:
    def __init__(self):
        self.log_manager = RichLogManager()
        self.logger = self.log_manager.get_logger()

    def docling_normalizer(self, path: Path):
        """
        Convert document to markdown with docling.

        Supported types:
        * .txt, .md
        * .pdf
        * .docx
        * .html
        * .csv
        * images (.jpg/.png/..)
        * spreadsheets (.xls/.xlsx)
        """

        from docling.document_converter import DocumentConverter

        # Initialize converter once (reusable)
        converter = DocumentConverter()

        self.logger.info(f"Processing: {path.name}")

        # Convert Document
        result = converter.convert(path)

        # Export to markdown
        return result.document.export_to_markdown()
    
    def normalize_document(self, path: Path):
        ext = path.suffix.lower()

        match ext:
            case ".txt" | ".md":
                return path.read_text(encoding="utf-8")

            case ".pdf":
                import fitz                    # PyMuPDF
                with fitz.open(str(path)) as pf:
                    return "\n".join(p.get_text() for p in pf)

            case ".docx":
                from docx import Document      # python-docx
                doc = Document(str(path))
                return "\n".join(p.text for p in doc.paragraphs)

            case ".html":
                from bs4 import BeautifulSoup  # BeautifulSoup4
                raw = path.read_text(encoding="utf-8")
                soup = BeautifulSoup(raw, "html.parser")
                return soup.get_text("\n")

            case ".csv":
                import pandas as pd            # pandas
                df = pd.read_csv(str(path))
                return df.to_string(index=False)

            case (".jpg", ".jpeg", ".png"):
                try:
                    from PIL import Image
                    import pytesseract          # tesseract OCR wrapper
                except ImportError:
                    raise HTTPException(
                        status_code=422,
                        detail="OCR dependency missing: install pillow and pytesseract",
                    )
                img = Image.open(str(path))
                return pytesseract.image_to_string(img)

            case (".xls", ".xlsx"):
                import openpyxl                 # openpyxl
                wb = openpyxl.load_workbook(str(path), data_only=True)
                sheets = []
                for sh in wb.sheetnames:
                    ws = wb[sh]
                    rows = [
                        "\t".join(
                            [str(v) if v is not None else "" for v in row]
                        )
                        for row in ws.iter_rows(values_only=True)
                    ]
                    sheets.append("\n".join(rows))
                return "\n--- Sheet Separator ---\n".join(sheets)

            case _:
                raise NotImplementedError(
                    f"Filetype '{ext}' is not yet supported"
                )